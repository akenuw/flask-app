import pandas as pd # type: ignore
from flask import Flask, render_template, request, redirect, url_for, session
from datetime import datetime
import openpyxl
import os

app = Flask(__name__)
app.secret_key = "your_secret_key_here"  # Used for session management

# Admin credentials (you can move these to a secure database later)
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "password123"

# Attendance time range
START_TIME = "05:00:00"
END_TIME = "08:00:00"

# Excel file to store attendance
attendance_file = "attendance.xlsx"

# Initialize Excel sheet if not already present
def initialize_excel():
    if not os.path.exists(attendance_file):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        # Add headers
        sheet.append(["Date", "Time", "Employee ID", "Name", "Status"])
        workbook.save(attendance_file)
        print("Attendance file created.")

initialize_excel()

# Employee database
employees = {
    1: "John Doe",
    2: "Jane Smith",
    3: "Alice Johnson",
    # Add more employees here...
}

# Login route
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        # Verify credentials
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session['logged_in'] = True
            return redirect(url_for('admin_dashboard'))
        else:
            return "Invalid credentials. Please try again."
    return render_template('login.html')

# Logout route
@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))
    
# Admin dashboard route
@app.route('/dashboard')
def admin_dashboard():
    if 'logged_in' in session and session['logged_in']:
        return render_template('dashboard.html')  # Admin dashboard template
    return redirect(url_for('login'))
    # Load attendance data
    workbook = openpyxl.load_workbook(attendance_file)
    sheet = workbook.active
    
    # Read all attendance records into a list
    attendance_records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        attendance_records.append(row)
    
    # Render the dashboard with records
    return render_template('dashboard.html', records=attendance_records)

# Route for scanning QR code
@app.route('/scan/<int:employee_id>')
def scan(employee_id):
    if employee_id not in employees:
        return "Error: Invalid QR code."

    # Get current time
    current_time = datetime.now().time()
    start_time = datetime.strptime(START_TIME, "%H:%M:%S").time()
    end_time = datetime.strptime(END_TIME, "%H:%M:%S").time()

    # Determine status
    if start_time <= current_time <= end_time:
        status = "Approved"
    else:
        status = "Error: Late or Outside Attendance Time"

    # Record attendance
    workbook = openpyxl.load_workbook(attendance_file)
    sheet = workbook.active
    sheet.append([datetime.now().date(), datetime.now().time(), employee_id, employees[employee_id], status])
    workbook.save(attendance_file)

    # Display status to the employee
    return f"<h1>{status}</h1>"

if __name__ == "__main__":
    app.run(debug=True)
